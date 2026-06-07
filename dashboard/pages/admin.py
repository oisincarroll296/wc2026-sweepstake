"""Admin page — password-protected event and draw controls."""
import sys
from pathlib import Path
sys.path.insert(0, str(Path(__file__).parent.parent.parent))

import streamlit as st
import pandas as pd

from dashboard.config import ADMIN_PASSWORD
from dashboard.components.ui import page_header, copyable_text

ROOT = Path(__file__).parent.parent.parent
DATA = ROOT / "data"

page_header("Admin", "Tournament management controls")

# ── Auth ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("---")
    st.markdown("**Admin Login**")

pwd = st.text_input("Password", type="password", placeholder="Enter admin password")
if not pwd:
    st.info("Enter the admin password to access controls.")
    st.stop()
if pwd != ADMIN_PASSWORD:
    st.error("Incorrect password.")
    st.stop()

st.success("Authenticated", icon="🔓")
st.divider()


def _refresh():
    st.cache_data.clear()


def _save_purchases(df: pd.DataFrame):
    df.to_csv(DATA / "purchases.csv", index=False)


def _save_statuses(df: pd.DataFrame):
    df.to_csv(DATA / "player_status.csv", index=False)


# ── Tabs ──────────────────────────────────────────────────────────────────
tabs = st.tabs([
    "Draw Events", "Purchases",
    "Locking", "Results Entry",
    "WhatsApp", "Draw Broadcast", "Deadlines",
])

# ─────────────────────────────────────────────
# Tab 0: Draw Events
# ─────────────────────────────────────────────
with tabs[0]:
    st.subheader("Run Draw Events")

    st.caption(
        "Use this panel to run the Initial Draw, Mulligan, Ninth Team, and Resurrection draws. "
        "Each draw is logged and can be broadcast via the Draw Broadcast tab."
    )

    event_type = st.selectbox("Event Type", [
        "INITIAL_DRAW", "MULLIGAN_DRAW", "GROUP_STAGE_CLOSE",
        "NINTH_TEAM_DRAW", "RESURRECTION_DRAW", "TOURNAMENT_COMPLETE",
    ])
    seed_input = st.text_input("Random Seed (leave blank for random)", placeholder="e.g. 42")
    seed = int(seed_input) if seed_input.strip().isdigit() else None

    if st.button(f"Run {event_type}", type="primary"):
        with st.spinner(f"Running {event_type}…"):
            try:
                from src.event_engine import run_event
                result = run_event(event_type, seed=seed)
                st.success(f"{event_type} executed successfully.")
                if "errors" in result and result["errors"]:
                    st.warning("Some players had errors:")
                    st.json(result["errors"])
                if "results" in result and result["results"]:
                    st.markdown("**Results:**")
                    st.json({k: str(v) for k, v in result["results"].items()})
                if "broadcast" in result:
                    st.markdown("**Broadcast text:**")
                    st.code(result["broadcast"], language=None)
                if "summary" in result:
                    st.info(result["summary"])
                _refresh()
            except Exception as exc:
                st.error(f"Error: {exc}")

# ─────────────────────────────────────────────
# Tab 1: Purchases
# ─────────────────────────────────────────────
with tabs[1]:
    st.subheader("Add Purchase")
    st.caption(
        "Buy In, Prediction Pack, and Insurance are processed immediately. "
        "Mulligan, Ninth Team, and Resurrection stay pending until their draw event runs."
    )

    with st.form("add_purchase"):
        from dashboard.data import get_participants
        players = get_participants() or []
        add_player = st.selectbox("Player", players or ["—"])
        add_type   = st.selectbox("Type", ["BUYIN", "PACK", "INSURANCE", "MULLIGAN", "NINTH", "RESURRECTION"])
        add_ref    = st.text_input("Payment Reference", placeholder="e.g. ALICE - BUY IN")
        add_sel    = st.text_input("Selection (Resurrection only)", placeholder="e.g. Spain->Germany")
        submitted  = st.form_submit_button("Add Purchase", type="primary")

        if submitted and add_player and add_player != "—":
            try:
                from src.competition import add_purchase, load_purchases, load_player_status
                from src.event_engine import process_pending_purchases

                p = load_purchases()
                s = load_player_status()
                p = add_purchase(add_player, add_type, add_ref, p, selection=add_sel)

                # Auto-process BUYIN / PACK / INSURANCE immediately
                up, us, msgs = process_pending_purchases(p, s)
                _save_purchases(up)
                _save_statuses(us)

                for m in msgs:
                    st.write(f"• {m}")
                st.success(f"{add_type} added for {add_player}.")
                _refresh()
            except Exception as exc:
                st.error(f"Error: {exc}")

    st.divider()

    # Current purchase log
    st.subheader("Purchase Log")
    from src.competition import load_purchases
    p = load_purchases()
    if p.empty:
        st.caption("No purchases recorded yet.")
    else:
        draw_pending = p[p["Status"] == "PENDING"]
        if not draw_pending.empty:
            st.info(
                f"{len(draw_pending)} draw-based purchase(s) pending — run their draw event to process them.",
                icon="⏳",
            )
        show = p[["Player", "PurchaseType", "Status", "Reference", "Timestamp"]].copy()
        show = show.sort_values("Timestamp", ascending=False)
        st.dataframe(show, use_container_width=True, hide_index=True)

# ─────────────────────────────────────────────
# Tab 2: Locking
# ─────────────────────────────────────────────
with tabs[2]:
    st.subheader("Lock Controls")

    from dashboard.data import is_predictions_locked
    pred_locked  = is_predictions_locked()
    buyin_locked = (DATA / "buyin_lock.txt").exists()

    col_status_a, col_status_b = st.columns(2)
    with col_status_a:
        if pred_locked:
            st.success("Predictions: LOCKED")
        else:
            st.warning("Predictions: Open")
    with col_status_b:
        if buyin_locked:
            st.success("Buy-ins: LOCKED")
        else:
            st.warning("Buy-ins: Open")

    st.divider()

    col_a, col_b = st.columns(2)

    with col_a:
        if not pred_locked:
            if st.button("Lock Predictions", type="primary"):
                try:
                    from src.competition import load_events, load_audit_log, load_predictions
                    from src.event_engine import lock_predictions
                    ev, log = lock_predictions(load_events(), load_audit_log())
                    ev.to_csv(DATA / "events.csv", index=False)
                    log.to_csv(DATA / "audit_log.csv", index=False)
                    preds = load_predictions()
                    n = len(preds) if not preds.empty else 0
                    st.success(f"Predictions locked. {n} player prediction(s) now public.")
                    _refresh()
                    st.rerun()
                except Exception as exc:
                    st.error(f"{exc}")
        else:
            st.info("Predictions are locked. To unlock, delete `data/predictions_lock.txt` and restart the app.")

    with col_b:
        if not buyin_locked:
            if st.button("Lock Buy-Ins", type="primary"):
                try:
                    from src.competition import load_events, load_audit_log, load_player_status
                    from src.event_engine import lock_buyins
                    s, ev, log = lock_buyins(load_player_status(), load_events(), load_audit_log())
                    ev.to_csv(DATA / "events.csv", index=False)
                    log.to_csv(DATA / "audit_log.csv", index=False)
                    paid = s[s["Status"] == "PAID"] if not s.empty else pd.DataFrame()
                    unpaid = s[s["Status"] != "PAID"] if not s.empty else pd.DataFrame()
                    st.success(f"Buy-ins locked. {len(paid)} paid / {len(unpaid)} unpaid.")
                    if not unpaid.empty:
                        st.warning("Unpaid players (excluded from prizes): " +
                                   ", ".join(unpaid["Player"].tolist()))
                    _refresh()
                    st.rerun()
                except Exception as exc:
                    st.error(f"{exc}")
        else:
            st.info("Buy-ins are locked. To unlock, delete `data/buyin_lock.txt` and restart the app.")

    st.divider()
    with st.expander("Unlock / Emergency Reset"):
        st.warning("Only use these if you made a mistake and the tournament has not started yet.")
        if st.button("Unlock Predictions"):
            p = DATA / "predictions_lock.txt"
            if p.exists():
                p.unlink()
                st.success("Predictions unlocked.")
                _refresh()
                st.rerun()
            else:
                st.info("Not locked.")
        if st.button("Unlock Buy-Ins"):
            p = DATA / "buyin_lock.txt"
            if p.exists():
                p.unlink()
                st.success("Buy-ins unlocked.")
                _refresh()
                st.rerun()
            else:
                st.info("Not locked.")

# ─────────────────────────────────────────────
# Tab 3: Results Entry
# ─────────────────────────────────────────────
with tabs[3]:
    result_mode = st.radio("Entry method", ["Single Team", "Upload Excel"], horizontal=True)

    if result_mode == "Single Team":
        st.subheader("Enter results for one team")
        st.caption(
            "Select a team and fill in their stats. Round Reached = the furthest round they reached. "
            "Teams still in the tournament should be left at their current round."
        )

        from dashboard.data import get_teams
        teams_df = get_teams()
        team_list = sorted(teams_df["Team"].tolist()) if not teams_df.empty else []

        with st.form("results_form"):
            res_team = st.selectbox("Team", team_list)

            # Pre-load existing values
            from src.scoring_engine import load_match_stats
            ms = load_match_stats()
            existing = {}
            if not ms.empty and res_team:
                row = ms[ms["Team"] == res_team]
                if not row.empty:
                    existing = row.iloc[0].to_dict()

            def _ev(col, default=0):
                v = existing.get(col, default)
                try:
                    return int(float(v)) if v != "" else default
                except Exception:
                    return default

            def _es(col):
                v = existing.get(col, "")
                return str(v) if v and str(v) != "nan" else ""

            r1, r2 = st.columns(2)
            with r1:
                st.markdown("**Group Stage**")
                g_goals  = st.number_input("Goals",         0, 50, _ev("GroupGoals"))
                g_cs     = st.number_input("Clean Sheets",  0, 10, _ev("GroupCleanSheets"))
                g_pw     = st.number_input("Penalty Wins",  0,  5, _ev("GroupPenaltyWins"))
                g_cw     = st.number_input("Comeback Wins", 0,  5, _ev("GroupComebackWins"))
                g_winner = st.checkbox("Group Winner", value=bool(_ev("GroupWinner")))
            with r2:
                st.markdown("**Knockout**")
                ko_goals = st.number_input("Goals",         0, 50, _ev("KnockoutGoals"))
                ko_cs    = st.number_input("Clean Sheets",  0, 10, _ev("KnockoutCleanSheets"))
                ko_pw    = st.number_input("Penalty Wins",  0,  5, _ev("KnockoutPenaltyWins"))
                ko_cw    = st.number_input("Comeback Wins", 0,  5, _ev("KnockoutComebackWins"))
                rounds   = ["", "GroupStage", "R16", "QF", "SF", "Final", "Winner"]
                cur_rnd  = _es("RoundReached")
                rnd_idx  = rounds.index(cur_rnd) if cur_rnd in rounds else 0
                rnd      = st.selectbox("Round Reached", rounds, index=rnd_idx)

            submitted_r = st.form_submit_button("Save Results", type="primary")
            if submitted_r and res_team:
                try:
                    from src.event_engine import update_results
                    ms = load_match_stats()
                    ms = update_results(res_team, {
                        "GroupGoals": g_goals, "GroupCleanSheets": g_cs,
                        "GroupPenaltyWins": g_pw, "GroupComebackWins": g_cw,
                        "GroupWinner": int(g_winner),
                        "KnockoutGoals": ko_goals, "KnockoutCleanSheets": ko_cs,
                        "KnockoutPenaltyWins": ko_pw, "KnockoutComebackWins": ko_cw,
                        "RoundReached": rnd,
                    }, ms)
                    ms.to_csv(DATA / "match_stats.csv", index=False)
                    st.success(f"Results saved for {res_team}.")
                    _refresh()
                except Exception as exc:
                    st.error(f"{exc}")

    else:
        st.subheader("Upload Results Excel")
        st.caption(
            "Upload the results_template.xlsx file. "
            "Generate a blank template first using: "
            "`.venv\\Scripts\\python.exe scripts/generate_results_excel.py`"
        )

        uploaded = st.file_uploader("Choose results_template.xlsx", type=["xlsx"])
        if uploaded:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(uploaded, data_only=True)

                from src.scoring_engine import load_match_stats
                ms = load_match_stats()

                gs_ws = wb["Group Stage"]
                ko_ws = wb["Knockout Rounds"]

                gs_headers = [c.value for c in next(gs_ws.iter_rows(min_row=1, max_row=1))]
                ko_headers = [c.value for c in next(ko_ws.iter_rows(min_row=1, max_row=1))]

                updated = 0
                for row in gs_ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]:
                        continue
                    team = str(row[0]).strip()
                    vals = dict(zip(gs_headers, row))
                    team_mask = ms["Team"] == team
                    if not team_mask.any():
                        continue
                    ms.loc[team_mask, "GroupGoals"]        = int(vals.get("Goals", 0) or 0)
                    ms.loc[team_mask, "GroupCleanSheets"]  = int(vals.get("Clean Sheets", 0) or 0)
                    ms.loc[team_mask, "GroupPenaltyWins"]  = int(vals.get("Penalty Wins", 0) or 0)
                    ms.loc[team_mask, "GroupComebackWins"] = int(vals.get("Comeback Wins", 0) or 0)
                    ms.loc[team_mask, "GroupWinner"]       = int(vals.get("Group Winner (1=Yes)", 0) or 0)
                    updated += 1

                for row in ko_ws.iter_rows(min_row=2, values_only=True):
                    if not row[0]:
                        continue
                    team = str(row[0]).strip()
                    vals = dict(zip(ko_headers, row))
                    team_mask = ms["Team"] == team
                    if not team_mask.any():
                        continue
                    ms.loc[team_mask, "KnockoutGoals"]        = int(vals.get("KO Goals", 0) or 0)
                    ms.loc[team_mask, "KnockoutCleanSheets"]  = int(vals.get("KO Clean Sheets", 0) or 0)
                    ms.loc[team_mask, "KnockoutPenaltyWins"]  = int(vals.get("KO Penalty Wins", 0) or 0)
                    ms.loc[team_mask, "KnockoutComebackWins"] = int(vals.get("KO Comeback Wins", 0) or 0)
                    rnd = str(vals.get("Round Reached", "") or "").strip()
                    if rnd:
                        ms.loc[team_mask, "RoundReached"] = rnd

                ms.to_csv(DATA / "match_stats.csv", index=False)
                st.success(f"Imported results for {updated} teams.")
                _refresh()
            except Exception as exc:
                st.error(f"Import failed: {exc}")

# ─────────────────────────────────────────────
# Tab 4: WhatsApp Update
# ─────────────────────────────────────────────
with tabs[4]:
    st.subheader("Generate WhatsApp Update")
    st.caption("Generates a formatted standings update to paste into your WhatsApp group.")

    if st.button("Generate Update", type="primary"):
        with st.spinner("Generating…"):
            try:
                from src.event_engine import generate_whatsapp_update
                from dashboard.data import (
                    get_prize_leaderboard, get_overall_leaderboard,
                    get_prize_pool, get_events, get_match_stats,
                )
                text = generate_whatsapp_update(
                    get_prize_leaderboard(), get_overall_leaderboard(),
                    get_prize_pool(), get_events(), get_match_stats(),
                )
                copyable_text("WhatsApp Update", text)
            except Exception as exc:
                st.error(f"{exc}")

# ─────────────────────────────────────────────
# Tab 5: Draw Broadcast
# ─────────────────────────────────────────────
with tabs[5]:
    st.subheader("Generate Draw Broadcast")
    st.caption("Generates a formatted draw announcement to paste into your WhatsApp group.")

    bc_type = st.selectbox("Draw Type", [
        "Initial Draw", "Mulligan Draw", "Ninth Team Draw", "Resurrection Draw",
    ])

    if st.button("Generate Broadcast", type="primary"):
        try:
            from src.event_engine import generate_draw_broadcast, load_allocation
            from src.competition import load_purchases

            results: dict[str, str] = {}

            if bc_type == "Initial Draw":
                alloc = load_allocation()
                if alloc.assignments:
                    results = {p: " | ".join(t) for p, t in alloc.assignments.items()}
                else:
                    st.warning(
                        "No draw found. Run INITIAL_DRAW first via the Draw Events tab, "
                        "then come back here to generate the broadcast."
                    )
                    st.stop()
            else:
                p = load_purchases()
                ptype_map = {
                    "Mulligan Draw":     "MULLIGAN",
                    "Ninth Team Draw":   "NINTH",
                    "Resurrection Draw": "RESURRECTION",
                }
                key = ptype_map[bc_type]
                done = p[(p["PurchaseType"] == key) & (p["Status"] == "PROCESSED")]
                if done.empty:
                    st.warning(f"No processed {bc_type} purchases found. Run the draw event first.")
                    st.stop()
                for _, row in done.iterrows():
                    results[str(row["Player"])] = str(row.get("Selection", "—"))

            text = generate_draw_broadcast(bc_type, results)
            copyable_text("Draw Broadcast", text)
        except Exception as exc:
            st.error(f"{exc}")

    st.divider()
    if st.button("Refresh All Scores"):
        _refresh()
        st.success("Cache cleared — scores will reload on next page view.")

# ─────────────────────────────────────────────
# Tab 6: Deadlines
# ─────────────────────────────────────────────
with tabs[6]:
    import json
    from datetime import datetime, timezone, date, time as dtime
    from dashboard.data import get_deadlines, save_deadlines, countdown, DEADLINE_LABELS

    st.subheader("Tournament Deadlines")
    st.caption(
        "Set the exact date and time for each deadline. Times are stored in UTC. "
        "The countdown shown on the Home page and Predictions Centre is derived from these values."
    )

    deadlines = get_deadlines()

    with st.form("deadlines_form"):
        updated: dict[str, str] = {}

        for key, label in DEADLINE_LABELS.items():
            iso = deadlines.get(key, "")
            try:
                dt = datetime.fromisoformat(iso).astimezone(timezone.utc)
                cur_date = dt.date()
                cur_time = dt.time().replace(second=0, microsecond=0)
            except Exception:
                cur_date = date(2026, 6, 11)
                cur_time = dtime(20, 0)

            cd = countdown(iso) if iso else "—"
            cd_text = f"  ·  **{cd}**" if cd not in ("—", "PASSED") else ("  ·  ~~passed~~" if cd == "PASSED" else "")

            st.markdown(f"**{label}**{cd_text}")
            col_d, col_t = st.columns([2, 1])
            with col_d:
                new_date = st.date_input(f"Date##{key}", value=cur_date, label_visibility="collapsed")
            with col_t:
                new_time = st.time_input(f"Time (UTC)##{key}", value=cur_time, label_visibility="collapsed", step=300)

            combined = datetime(
                new_date.year, new_date.month, new_date.day,
                new_time.hour, new_time.minute, 0,
                tzinfo=timezone.utc,
            )
            updated[key] = combined.isoformat()
            st.markdown("")

        if st.form_submit_button("Save All Deadlines", type="primary"):
            save_deadlines(updated)
            _refresh()
            st.success("Deadlines saved.")
            st.rerun()
